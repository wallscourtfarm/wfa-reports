"""
Pupil Voice importer for WFA Reports Manager.
Reads a Microsoft Forms export Excel, cleans and narrativises each
pupil's answers using Claude, and outputs a two-column Excel.
"""
import io
import time
import anthropic
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Forms column names (as they appear after header row)
COL_NAME       = "Select Name"
COL_ENQ        = "The enquiry which I enjoyed the most this year was"
COL_ENQ_WHY   = "because"
COL_SOB        = "The state of being I enjoy the most is"
COL_SOB_WHY   = "because 1"
COL_CONF       = "I want to become more confident with"
COL_CONF_WHY  = "because 2"
COL_Y5         = "In Year 5 I am looking forward to"

SYSTEM_PROMPT = """You are a teaching assistant helping tidy up Year 4 pupil (aged 8-9) responses to a school report questionnaire. 

Your job is to take their raw answers and produce a single, clean narrative paragraph written in the child's own voice. 

Rules:
- Correct all spelling, grammar and punctuation errors
- Keep the child's meaning and sentiment exactly — do not add or invent anything
- Write in first person ("I enjoyed...", "I want to...", "I am looking forward to...")
- Weave all the answers into natural flowing prose — do not list them as separate sentences if they can join naturally
- Include the sense of each question implicitly (e.g. "The enquiry I enjoyed most was..." should be the opening)
- Do not use bullet points or any formatting — plain prose only
- Keep it natural and age-appropriate — this is a Year 4 child speaking
- Output ONLY the narrative paragraph, nothing else — no preamble, no explanation"""


def _build_prompt(row: dict) -> str:
    return f"""Here are a Year 4 pupil's answers to a report questionnaire. Clean them up and write a single narrative paragraph in their voice.

Enquiry enjoyed most: {row.get(COL_ENQ, '')}
Because: {row.get(COL_ENQ_WHY, '')}
State of being enjoyed most: {row.get(COL_SOB, '')}
Because: {row.get(COL_SOB_WHY, '')}
Want to be more confident with: {row.get(COL_CONF, '')}
Because: {row.get(COL_CONF_WHY, '')}
In Year 5 looking forward to: {row.get(COL_Y5, '')}"""


def _clean_narrative(row: dict) -> str:
    client = anthropic.Anthropic(api_key=st.secrets["ANTHROPIC_API_KEY"])
    msg = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=400,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": _build_prompt(row)}],
    )
    return msg.content[0].text.strip()


def parse_forms_export(file_bytes: bytes) -> pd.DataFrame:
    """Read the Forms Excel and return a tidy DataFrame."""
    df = pd.read_excel(io.BytesIO(file_bytes))
    # Strip non-breaking spaces and whitespace from column names (Forms export quirk)
    df.columns = [str(c).replace("\xa0", " ").strip() for c in df.columns]
    # Keep only rows that have a name
    df = df[df[COL_NAME].notna() & (df[COL_NAME].astype(str).str.strip() != "")]
    # Fill NaN answers with empty string
    answer_cols = [COL_ENQ, COL_ENQ_WHY, COL_SOB, COL_SOB_WHY,
                   COL_CONF, COL_CONF_WHY, COL_Y5]
    for col in answer_cols:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str)
    return df.reset_index(drop=True)


def build_output_excel(results: list[dict]) -> bytes:
    """
    Build a two-column Excel: Last Name | First Name | Pupil Voice.
    Sorted alphabetically by last name.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Pupil Voice"

    # Styles
    BLUE = "1798D3"
    header_fill = PatternFill("solid", fgColor=BLUE)
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    body_font   = Font(name="Calibri", size=11)
    wrap_align  = Alignment(wrap_text=True, vertical="top")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    for col, header in enumerate(["Name", "Pupil Voice"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    # Sort by last name if possible, otherwise by full name
    def sort_key(r):
        name = r.get("name", "")
        parts = name.strip().rsplit(" ", 1)
        return parts[-1].lower() if len(parts) > 1 else name.lower()

    for row_idx, rec in enumerate(sorted(results, key=sort_key), start=2):
        name_cell = ws.cell(row=row_idx, column=1, value=rec["name"])
        pv_cell   = ws.cell(row=row_idx, column=2, value=rec["pupil_voice"])
        for cell in [name_cell, pv_cell]:
            cell.font      = body_font
            cell.border    = border
            cell.alignment = wrap_align
        # Auto row height — rough estimate
        lines = max(1, len(rec["pupil_voice"]) // 95 + rec["pupil_voice"].count("\n"))
        ws.row_dimensions[row_idx].height = max(30, lines * 15)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def process_forms_export(file_bytes: bytes, progress_cb=None) -> tuple[list[dict], list[str]]:
    """
    Main entry point. Returns (results, errors).
    results: list of {name, pupil_voice}
    errors:  list of name strings that failed
    """
    df = parse_forms_export(file_bytes)
    results, errors = [], []

    for i, (_, row) in enumerate(df.iterrows()):
        name = str(row[COL_NAME]).strip()
        try:
            narrative = _clean_narrative(row.to_dict())
            results.append({"name": name, "pupil_voice": narrative})
        except Exception as e:
            errors.append(f"{name}: {e}")
        if progress_cb:
            progress_cb(i + 1, len(df), name)
        time.sleep(0.2)  # avoid rate limiting

    return results, errors
