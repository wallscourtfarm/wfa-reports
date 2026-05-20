"""
Export class data to report-data.xlsx format.
ZIP-level surgery: copies the template verbatim, then:
  - replaces Data sheet XML with fresh openpyxl-generated content
  - replaces styles.xml with fresh one so style IDs are consistent
  - strips <drawing> element from Set up Report sheet XML
  - strips calcChain.xml (stale after row changes, Excel rebuilds it)
"""
import io
import os
import re
import zipfile
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Grade mapping ──────────────────────────────────────────────────────────────
GRADE_MAP = {
    'D':'D', 'GD':'D', 'O':'O', 'O1':'O', 'O2':'O', 'Y':'Y',
    'A - Y2':'A Y2', 'A - Y3':'A Y3', 'A - Y4':'A Y4', 'A - Y5':'A Y5',
}

def _grade(v): return GRADE_MAP.get(str(v or '').strip(), str(v or ''))


# ── Formulas for columns I and K ───────────────────────────────────────────────
def _att_formula(row):
    r = str(row)
    return (f"=IF(H{r}>='Set up Report'!$B$46,\"A\","
            f"IF(H{r}>='Set up Report'!$B$47,\"B\","
            f"IF(H{r}<'Set up Report'!$B$49,\"D\",\"C\")))")

def _punc_formula(row):
    r = str(row)
    return (f"=IF(AND(J{r}='Set up Report'!$C$46),\"A\","
            f"IF(AND(J{r}<='Set up Report'!$C$47),\"B\","
            f"IF(AND(J{r}<='Set up Report'!$C$48),\"C\","
            f"IF(AND(J{r}>'Set up Report'!$C$48),\"D\"))))")


# ── Locate the template ────────────────────────────────────────────────────────
def _template_path():
    sd = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(sd, 'data', 'static', 'report_template.xlsx'),
        os.path.join('/mount/src/wfa-reports/data/static', 'report_template.xlsx'),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


# ── Styles ─────────────────────────────────────────────────────────────────────
WRAP  = Alignment(wrap_text=True, vertical='top')
TOP   = Alignment(vertical='top')
THIN  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)
BODY  = Font(name='Calibri', size=10, bold=False)
HDR   = Font(name='Calibri', size=10, bold=True)
WIDTHS = [16,14,18,24,8,8,8,14,18,20,18,60,60,60,60,60,60]


# ── Build workbook with Data sheet only ───────────────────────────────────────
def _build_data_workbook(pupils):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Row 1 notes
    ws.cell(1, 2).value = 'Copy Paste from DOOYA\t\t'
    ws.cell(1, 5).value = 'Copy Paste from DOOYA'
    ws.cell(1, 8).value = 'copy-paste from Welcome Zone Report'
    ws.cell(1, 10).value = 'copy-paste from Welcome Zone Report'

    # Row 2 headers
    headers = [
        'ID', 'First Name', 'Last Name', 'Full Name',
        'R', 'W', 'M',
        'Attendance', 'Att Code', 'Punctuality (Lates) #', 'Punc Code',
        'Reader Teacher comments', 'Writer Teacher comments',
        'Mathematician Teacher comments', '21st C learner Teacher comments',
        'Rights and Responsibilities Teacher comments', 'Pupil Voice',
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, value=h)
        c.font = HDR
        c.border = THIN

    # Data rows
    for ri, pupil in enumerate(pupils, start=3):
        fn    = (pupil.get('first_name') or '').strip()
        ln    = (pupil.get('last_name')  or '').strip()
        grades = pupil.get('grades', {})
        coms  = pupil.get('comments', {})
        att   = str(pupil.get('attendance')  or '').strip()
        lates = str(pupil.get('punctuality') or '').strip()

        try:    lates_v = int(lates)
        except: lates_v = None if not lates else lates

        row = {
            1:  f"ch{ri-2:02d}",
            2:  fn,
            3:  ln,
            4:  f"{fn} {ln}".strip(),
            5:  _grade(grades.get('R','')),
            6:  _grade(grades.get('W','')),
            7:  _grade(grades.get('M','')),
            8:  float(att) if att else None,
            9:  _att_formula(ri),
            10: lates_v,
            11: _punc_formula(ri),
            12: coms.get('reader','')        or '',
            13: coms.get('writer','')        or '',
            14: coms.get('mathematician','') or '',
            15: coms.get('learner_21c','')   or '',
            16: coms.get('rights','')        or '',
            17: (pupil.get('pupil_voice') or '').strip(),
        }

        for col, val in row.items():
            c = ws.cell(row=ri, column=col, value=val)
            c.font      = BODY
            c.border    = THIN
            c.alignment = WRAP if col >= 12 else TOP

        ws.row_dimensions[ri].height = 80

    for ci, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A3'
    return wb


# ── Main export ────────────────────────────────────────────────────────────────
def export_excel(class_data: dict, settings: dict = None) -> io.BytesIO:
    settings = settings or {}
    pupils = sorted(
        class_data.get('pupils', []),
        key=lambda p: (p.get('last_name',''), p.get('first_name',''))
    )

    tpl = _template_path()
    if not tpl:
        wb = _build_data_workbook(pupils)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # Save the fresh workbook and extract what we need from it
    fresh_wb = _build_data_workbook(pupils)
    fresh_buf = io.BytesIO()
    fresh_wb.save(fresh_buf)
    fresh_buf.seek(0)
    with zipfile.ZipFile(fresh_buf) as fz:
        fresh_data_xml  = fz.read('xl/worksheets/sheet1.xml')
        fresh_styles_xml = fz.read('xl/styles.xml')
        fresh_strings    = fz.read('xl/sharedStrings.xml') if 'xl/sharedStrings.xml' in fz.namelist() else None

    # Find which entry is the Data sheet in the template
    with open(tpl, 'rb') as f:
        tpl_bytes = f.read()

    with zipfile.ZipFile(io.BytesIO(tpl_bytes)) as tz:
        wb_xml   = tz.read('xl/workbook.xml').decode()
        wb_rels  = tz.read('xl/_rels/workbook.xml.rels').decode()

        m = re.search(r'<sheet[^>]+name=["\']Data["\'][^>]+r:id=["\']([\w]+)["\'][^>]*/>', wb_xml)
        if not m:
            m = re.search(r'<sheet[^>]+r:id=["\']([\w]+)["\'][^>]+name=["\']Data["\'][^>]*/>', wb_xml)
        rid = m.group(1) if m else 'rId2'

        rel_m = re.search(
            r'<Relationship[^>]+Id=["\'' + re.escape(rid) + r'["\'][^>]+Target=["\']([^"\']+ )["\'][^>]*/>',
            wb_rels
        )
        if rel_m:
            target = rel_m.group(1).strip()
            data_path = 'xl/' + target.lstrip('./')
        else:
            data_path = 'xl/worksheets/sheet2.xml'

        # Build output
        out = io.BytesIO()
        SKIP = {'xl/calcChain.xml'}  # stale — Excel will rebuild

        with zipfile.ZipFile(out, 'w', compression=zipfile.ZIP_DEFLATED) as oz:
            for item in tz.infolist():
                if item.filename in SKIP:
                    continue

                if item.filename == data_path:
                    oz.writestr(item.filename, fresh_data_xml)

                elif item.filename == 'xl/styles.xml':
                    oz.writestr(item.filename, fresh_styles_xml)

                elif item.filename == 'xl/sharedStrings.xml' and fresh_strings:
                    oz.writestr(item.filename, fresh_strings)

                else:
                    raw = tz.read(item.filename)

                    # Strip <drawing.../> from Set up Report sheet XML
                    if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                        raw = re.sub(rb'<drawing[^/]*/>', b'', raw)

                    oz.writestr(item.filename, raw)

    out.seek(0)
    return out
